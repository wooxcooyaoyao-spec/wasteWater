"""
污泥处理数据 Excel 处理模块

功能：
1. 读取 Excel 数据并解析
2. 基于参数调整进行实时计算
3. 生成分析报告
4. 支持 xlwings 集成（可选）
"""

from pathlib import Path
from typing import Dict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from wastewater_treatment_calc import WastewaterCalculator


class ExcelDataHandler:
    """Excel 数据处理类"""

    def __init__(self, excel_path: str = None):
        """
        初始化处理器

        Args:
            excel_path: Excel 文件路径
        """
        self.excel_path = excel_path
        self.df = None
        self.calculator = WastewaterCalculator(area=1.0)

        if excel_path and Path(excel_path).exists():
            self.load_excel(excel_path)

    def load_excel(self, excel_path: str) -> list:
        """
        加载 Excel 文件

        Args:
            excel_path: Excel 文件路径

        Returns:
            列表形式的数据
        """
        wb = load_workbook(excel_path)
        ws = wb.active
        self.df = []
        for row in ws.iter_rows(values_only=True):
            self.df.append(row)
        self.excel_path = excel_path
        print(f"✓ 加载 Excel 文件: {excel_path}")
        return self.df

    def parse_mlss_table(self) -> Dict:
        """
        解析标准的 MLSS 浓度表

        Returns:
            包含解析后的表格结构和数据的字典
        """
        if self.df is None:
            return {'error': '未加载 Excel 文件'}

        # 假设第一行是标题（MLSS 值）
        # 假设第二行是副标题
        # 从第三行开始是数据

        mlss_values = []
        equivalent_values = []
        slr_data = []

        try:
            # 获取 MLSS 列标题（从第一行第二列开始）
            if len(self.df) > 0 and len(self.df[0]) > 1:
                for col_idx in range(1, len(self.df[0])):
                    try:
                        mlss_val = float(self.df[0][col_idx])
                        mlss_values.append(mlss_val)
                    except:
                        pass

            # 获取数据行（从第三行开始）
            for row_idx in range(2, len(self.df)):
                try:
                    eq_val = float(self.df[row_idx][0])
                    equivalent_values.append(eq_val)

                    row_data = []
                    for col_idx in range(1, len(mlss_values) + 1):
                        if col_idx < len(self.df[row_idx]):
                            slr_val = float(self.df[row_idx][col_idx])
                            row_data.append(slr_val)
                    if row_data:
                        slr_data.append(row_data)
                except:
                    continue

            return {
                'mlss_values': mlss_values,
                'equivalent_values': equivalent_values,
                'slr_data': slr_data,
                'shape': (len(equivalent_values), len(mlss_values))
            }
        except Exception as e:
            return {'error': str(e)}

    def generate_analysis_report(self, output_file: str = None) -> list:
        """
        生成分析报告：计算每个点的运行状态

        Args:
            output_file: 输出文件路径（可选）

        Returns:
            列表格式的报告
        """
        table_info = self.parse_mlss_table()
        if 'error' in table_info:
            return []

        results = []

        for eq in table_info['equivalent_values']:
            for mlss_idx, mlss in enumerate(table_info['mlss_values']):
                eq_idx = table_info['equivalent_values'].index(eq)
                if eq_idx < len(table_info['slr_data']) and mlss_idx < len(table_info['slr_data'][eq_idx]):
                    slr = table_info['slr_data'][eq_idx][mlss_idx]

                    # 验证状态
                    check = self.calculator.check_operating_point(mlss, eq)

                    results.append({
                        'MLSS (mg/L)': int(mlss),
                        'Equivalent (L/s)': int(eq),
                        'SLR (kg/h/m²)': f'{slr:.2f}',
                        'MLSS Status': check['mlss']['status'],
                        'Flow Status': check['equivalent_flow']['status'],
                        'SLR Status': check['slr']['status'],
                        'Overall Safe': '✓' if check['overall_safe'] else '✗',
                    })

        if output_file:
            # 保存为 Excel
            wb = Workbook()
            ws = wb.active

            # 写入表头
            headers = list(results[0].keys()) if results else []
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # 写入数据
            for row_idx, item in enumerate(results, start=2):
                for col_idx, value in enumerate(item.values(), start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(output_file)
            print(f"✓ 分析报告已保存: {output_file}")

        return results

    def create_comparison_excel(self, output_file: str, variations: Dict) -> None:
        """
        创建对比分析 Excel

        Args:
            output_file: 输出文件路径
            variations: 参数变化字典，如：
                {
                    '基准': {'mlss': 3500, 'flow': 100},
                    '高流量': {'mlss': 3500, 'flow': 120},
                    '高浓度': {'mlss': 4000, 'flow': 100},
                }
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "参数对比"

        # 表头
        headers = ['场景', 'MLSS (mg/L)', 'Equivalent (L/s)', 'SLR (kg/h/m²)',
                   'MLSS 状态', 'Flow 状态', 'SLR 状态', '整体安全']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 数据行
        for row_idx, (scenario, params) in enumerate(variations.items(), start=2):
            mlss = params['mlss']
            flow = params['flow']
            check = self.calculator.check_operating_point(mlss, flow)

            ws.cell(row=row_idx, column=1, value=scenario)
            ws.cell(row=row_idx, column=2, value=mlss)
            ws.cell(row=row_idx, column=3, value=flow)
            ws.cell(row=row_idx, column=4, value=round(check['calculated_slr'], 2))
            ws.cell(row=row_idx, column=5, value=check['mlss']['status'])
            ws.cell(row=row_idx, column=6, value=check['equivalent_flow']['status'])
            ws.cell(row=row_idx, column=7, value=check['slr']['status'])
            ws.cell(row=row_idx, column=8, value='✓' if check['overall_safe'] else '✗')

            # 条件格式
            for col in range(1, 9):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = Alignment(horizontal="center")
                if col == 8:  # 整体安全列
                    if check['overall_safe']:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # 调整列宽
        ws.column_dimensions['A'].width = 15
        for col in range(2, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18

        wb.save(output_file)
        print(f"✓ 对比分析 Excel 已保存: {output_file}")

    def create_sensitivity_analysis(self, output_file: str, base_mlss: float = 3500,
                                    base_flow: float = 100) -> None:
        """
        创建敏感性分析 - 显示参数变化对 SLR 的影响

        Args:
            output_file: 输出文件路径
            base_mlss: 基准 MLSS (mg/L)
            base_flow: 基准流量 (L/s)
        """
        wb = Workbook()

        # Sheet 1: MLSS 变化影响
        ws1 = wb.active
        ws1.title = "MLSS变化影响"

        mlss_variations = range(2000, 5500, 500)
        ws1['A1'] = 'MLSS 敏感性分析'
        ws1['A1'].font = Font(bold=True, size=14)

        ws1['A3'] = 'MLSS (mg/L)'
        ws1['B3'] = 'SLR (kg/h/m²)'
        ws1['C3'] = '相对变化 (%)'

        base_slr = self.calculator.calculate_slr(base_mlss, base_flow)

        for row_idx, mlss in enumerate(mlss_variations, start=4):
            slr = self.calculator.calculate_slr(mlss, base_flow)
            change = ((slr - base_slr) / base_slr) * 100

            ws1.cell(row=row_idx, column=1, value=mlss)
            ws1.cell(row=row_idx, column=2, value=round(slr, 2))
            ws1.cell(row=row_idx, column=3, value=round(change, 2))

        # Sheet 2: 流量变化影响
        ws2 = wb.create_sheet("流量变化影响")

        flow_variations = range(60, 175, 10)
        ws2['A1'] = '流量敏感性分析'
        ws2['A1'].font = Font(bold=True, size=14)

        ws2['A3'] = 'Equivalent (L/s)'
        ws2['B3'] = 'SLR (kg/h/m²)'
        ws2['C3'] = '相对变化 (%)'

        for row_idx, flow in enumerate(flow_variations, start=4):
            slr = self.calculator.calculate_slr(base_mlss, flow)
            change = ((slr - base_slr) / base_slr) * 100

            ws2.cell(row=row_idx, column=1, value=flow)
            ws2.cell(row=row_idx, column=2, value=round(slr, 2))
            ws2.cell(row=row_idx, column=3, value=round(change, 2))

        # 格式化
        for ws in [ws1, ws2]:
            for row in range(3, 20):
                for col in range(1, 4):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

        wb.save(output_file)
        print(f"✓ 敏感性分析 Excel 已保存: {output_file}")


# 使用示例
def example_excel_operations(data_dir: str = None, output_dir: str = None):
    """
    演示 Excel 处理示例

    Args:
        data_dir: 数据目录（默认为当前目录下的 data 文件夹）
        output_dir: 输出目录（默认为当前目录下的 output 文件夹）
    """
    print("\n" + "=" * 70)
    print("Excel 数据处理示例")
    print("=" * 70)

    # 使用相对路径
    if data_dir is None:
        data_dir = Path(__file__).parent / "data"
    else:
        data_dir = Path(data_dir)

    if output_dir is None:
        output_dir = Path(__file__).parent / "output"
    else:
        output_dir = Path(output_dir)

    # 查找 Excel 文件
    excel_file = data_dir / "MLSS浓度表.xlsx"

    if excel_file.exists():
        handler = ExcelDataHandler(str(excel_file))

        # 生成对比分析
        variations = {
            '基准运行': {'mlss': 3500, 'flow': 100},
            '高流量运行': {'mlss': 3500, 'flow': 120},
            '高浓度运行': {'mlss': 4000, 'flow': 100},
            '超高负荷': {'mlss': 4500, 'flow': 150},
        }

        output_dir.mkdir(parents=True, exist_ok=True)

        handler.create_comparison_excel(
            str(output_dir / "参数对比分析.xlsx"),
            variations
        )

        # 生成敏感性分析
        handler.create_sensitivity_analysis(
            str(output_dir / "敏感性分析.xlsx"),
            base_mlss=3500,
            base_flow=100
        )

        print("\n✓ 所有分析文件已生成！")
        print(f"  数据目录: {data_dir.resolve()}")
        print(f"  输出目录: {output_dir.resolve()}")
    else:
        print(f"✗ 未找到 Excel 文件: {excel_file.resolve()}")
        print(f"  请确保文件存在于: {data_dir.resolve()}")


if __name__ == '__main__':
    example_excel_operations()

